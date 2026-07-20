#!/usr/bin/env python3
"""Audit numeric laboratory coverage in the pinned diagnostic benchmarks.

The scanner is intentionally catalog-first: it counts a numeric measurement
only when a curated lab alias is followed by a value, then reports unit
variants and whether the runtime can safely compare them. A conservative
generic pass surfaces possible uncatalogued laboratory labels for human review.
No case narrative is copied into the committed outputs.
"""

from __future__ import annotations

import argparse
import collections
import hashlib
import json
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from agentclinic_tree_dx.knowledge.finding_normalizer import FindingNormalizer  # noqa: E402


DEFAULT_DATASET_MANIFEST = ROOT / "data" / "eval" / "lab_reference_dataset_manifest.json"
DEFAULT_SOURCE_MANIFEST = ROOT / "data" / "knowledge_raw" / "lab_reference_sources.json"
DEFAULT_EXTENSION = ROOT / "data" / "knowledge_raw" / "lab_reference_range_extensions.json"
DEFAULT_CATALOG = ROOT / "data" / "knowledge_raw" / "lab_reference_ranges.json"
DEFAULT_LOINC = ROOT / "data" / "knowledge_raw" / "loinc2hpo_annotations.json"
DEFAULT_CONVERSIONS = ROOT / "data" / "knowledge_raw" / "unit_conversions.json"
DEFAULT_DATASET_DIR = ROOT / ".cache" / "lab-reference-audit"
DEFAULT_JSON = ROOT / "data" / "eval" / "lab_reference_coverage_report.json"
DEFAULT_MARKDOWN = ROOT / "LAB_REFERENCE_RANGE_DATA_SOURCES.md"


VALUE_PREFIX_RE = re.compile(
    r"^(?:\s|:|=|–|-)*"
    r"(?:(?:level|count|concentration|activity|value|result|was|is|of|at|measured|"
    r"increased|decreased|rose|fell|from|to|approximately|about)\s*){0,6}"
    r"(?P<qualifier><=?|>=?|≤|≥)?\s*(?P<value>\d[\d,]*(?:\.\d+)?)",
    re.IGNORECASE,
)

UNIT_AFTER_RE = re.compile(
    r"^\s*(?P<unit>"
    r"(?:[x×]\s*10\s*\^?\s*\d+\s*/\s*(?:L|[μµu]L)|"
    r"10\s*\^\s*\d+\s*/\s*(?:L|[μµu]L)|"
    r"(?:cells|K|M|mil|thou)?\s*/\s*(?:mm3|[μµu]L)|"
    r"(?:mEq|mmol|μmol|µmol|umol|nmol|pmol)\s*/\s*L|"
    r"(?:kg|g|mg|μg|µg|ug|mcg|ng|pg)\s*/\s*(?:dL|mL|L|24\s*h)|"
    r"(?:U|IU|mIU|μIU|µIU|kIU)\s*/\s*(?:mL|L)|"
    r"mL\s*/\s*min(?:\s*/\s*1\.73\s*m2)?|"
    r"mOsm(?:ol)?\s*/\s*kg(?:\s*H2O)?|"
    r"mm\s*Hg|kPa|mm\s*/\s*h(?:r)?|seconds?|secs?|sec|s|fL|pg|%|\{ratio\}|\{pH\})"
    r"(?:\s*(?:FEU|DDU))?)",
    re.IGNORECASE,
)

GENERIC_MEASUREMENT_RE = re.compile(
    r"(?P<label>[A-Za-z][A-Za-z0-9+α-ωΑ-Ω/'().-]*(?:\s+[A-Za-z][A-Za-z0-9+α-ωΑ-Ω/'().-]*){0,5})"
    r"\s*[:=]\s*(?P<qualifier><=?|>=?|≤|≥)?\s*"
    r"(?P<value>\d[\d,]*(?:\.\d+)?)\s*"
    + UNIT_AFTER_RE.pattern.removeprefix("^\\s*"),
    re.IGNORECASE,
)

LAB_HINT_RE = re.compile(
    r"(?:blood|serum|plasma|urine|csf|count|level|protein|antigen|antibody|"
    r"globulin|complement|troponin|peptide|hormone|enzyme|kinase|transferase|"
    r"cytokine|interleukin|factor|ratio|saturation|acid|glucose|sodium|potassium|"
    r"calcium|chloride|phosphate|creatinine|urea|bilirubin|hemoglobin|platelet|"
    r"\bil-\d|\bifn|\btnf|\bca\s*\d)",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class TextRecord:
    dataset_id: str
    row_id: int
    text: str


@dataclass(frozen=True)
class Measurement:
    canonical: str
    alias: str
    value: float
    qualifier: str
    unit: str
    start: int
    end: int


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _verify_artifact(path: Path, record: dict) -> None:
    if not path.is_file():
        raise FileNotFoundError(
            f"{record['id']} is missing at {path}; run scripts/download_lab_audit_datasets.py"
        )
    if path.stat().st_size != record["bytes"] or _sha256(path) != record["sha256"]:
        raise ValueError(f"{record['id']} does not match the pinned size/SHA-256")


def _load_parquet(path: Path, field: str) -> list[str]:
    try:
        import pyarrow.parquet as pq
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("Install audit dependencies with: pip install -e '.[lab-audit]'") from exc
    table = pq.read_table(path, columns=[field])
    return [str(value or "") for value in table.column(field).to_pylist()]


def _load_xlsx(path: Path, field: str) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("Install audit dependencies with: pip install -e '.[lab-audit]'") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))
    if field not in header:
        raise KeyError(f"{field!r} is absent from {path}")
    index = header.index(field)
    return [
        str(row[index] or "")
        for row in rows
        if any(value is not None for value in row)
    ]


def _inspect_zip_jsonl(path: Path) -> dict:
    rows = 0
    keys: set[str] = set()
    numeric_values = 0
    with zipfile.ZipFile(path) as archive:
        members = sorted(name for name in archive.namelist() if name.endswith(".jsonl"))
        for member in members:
            with archive.open(member) as handle:
                for raw_line in handle:
                    if not raw_line.strip():
                        continue
                    rows += 1
                    record = json.loads(raw_line)
                    keys.update(record)
                    numeric_values += sum(
                        isinstance(value, (int, float)) and not isinstance(value, bool)
                        for value in record.values()
                    )
    return {"rows": rows, "members": members, "fields": sorted(keys), "numeric_scalar_fields": numeric_values}


def iter_dataset_texts(dataset_dir: Path, manifest: dict) -> tuple[list[TextRecord], dict[str, dict]]:
    records: list[TextRecord] = []
    schema_checks: dict[str, dict] = {}
    for dataset in manifest["datasets"]:
        path = dataset_dir / dataset["relative_path"]
        _verify_artifact(path, dataset)
        if not dataset.get("numeric_lab_audit"):
            if dataset["format"] == "zip_jsonl":
                schema_checks[dataset["id"]] = _inspect_zip_jsonl(path)
            continue
        if dataset["format"] == "parquet":
            texts = _load_parquet(path, dataset["text_field"])
        elif dataset["format"] == "xlsx":
            texts = _load_xlsx(path, dataset["text_field"])
        else:  # pragma: no cover - manifest guard
            raise ValueError(f"unsupported format: {dataset['format']}")
        if len(texts) != dataset["rows"]:
            raise ValueError(f"{dataset['id']}: expected {dataset['rows']} rows, found {len(texts)}")
        records.extend(TextRecord(dataset["id"], i, text) for i, text in enumerate(texts))
    return records, schema_checks


def _alias_patterns(catalog: dict) -> list[tuple[str, str, re.Pattern]]:
    patterns: list[tuple[str, str, re.Pattern]] = []
    for canonical, entry in catalog.items():
        aliases = {canonical.replace("_", " "), *entry.get("aliases", [])}
        for alias in aliases:
            alias = alias.strip()
            if len(alias) < 2:
                continue
            pattern = re.compile(
                rf"(?<![A-Za-z0-9]){re.escape(alias)}(?![A-Za-z0-9])",
                re.IGNORECASE,
            )
            patterns.append((canonical, alias, pattern))
    return sorted(patterns, key=lambda item: len(item[1]), reverse=True)


def scan_measurements(text: str, patterns: list[tuple[str, str, re.Pattern]]) -> list[Measurement]:
    candidates: list[Measurement] = []
    for canonical, alias, pattern in patterns:
        for alias_match in pattern.finditer(text):
            tail = text[alias_match.end() : alias_match.end() + 100]
            value_match = VALUE_PREFIX_RE.match(tail)
            if value_match is None:
                continue
            value_end = alias_match.end() + value_match.end()
            unit_match = UNIT_AFTER_RE.match(text[value_end : value_end + 60])
            unit = (unit_match.group("unit") if unit_match else "").strip()
            end = value_end + (unit_match.end() if unit_match else 0)
            candidates.append(
                Measurement(
                    canonical=canonical,
                    alias=alias,
                    value=float(value_match.group("value").replace(",", "")),
                    qualifier=value_match.group("qualifier") or "",
                    unit=unit,
                    start=alias_match.start(),
                    end=end,
                )
            )

    # Prefer the longest alias when nested aliases point at the same value
    # ("high-sensitivity troponin I" must not also become "troponin I").
    accepted: list[Measurement] = []
    occupied: list[tuple[int, int]] = []
    for measurement in sorted(candidates, key=lambda m: (-len(m.alias), m.start, m.end)):
        if any(measurement.start < end and measurement.end > start for start, end in occupied):
            continue
        occupied.append((measurement.start, measurement.end))
        accepted.append(measurement)
    return sorted(accepted, key=lambda m: m.start)


def _unit_status(normalizer: FindingNormalizer, catalog: dict, measurement: Measurement) -> str:
    if not measurement.unit:
        return "unit_missing"
    entry = catalog[measurement.canonical]
    ranges = entry.get("reference_ranges", [])
    if not ranges:
        return "context_required"
    observed = normalizer._normalize_unit(measurement.unit)
    for ref in ranges:
        target = normalizer._normalize_unit(str(ref.get("unit", "")))
        if observed == target:
            return "supported"
        if normalizer._convert_unit(measurement.canonical, measurement.value, observed, target) is not None:
            return "supported"
    return "unsupported_unit"


def _reference_sources(entry: dict) -> list[str]:
    sources = {
        record.get("source")
        for kind in ("reference_ranges", "decision_limits")
        for record in entry.get(kind, [])
        if record.get("source")
    }
    sources.update(entry.get("interpretation_sources", []))
    return sorted(sources)


def _generic_candidates(text: str, known_spans: list[tuple[int, int]]) -> list[tuple[str, float, str]]:
    output: list[tuple[str, float, str]] = []
    for match in GENERIC_MEASUREMENT_RE.finditer(text):
        if any(match.start() < end and match.end() > start for start, end in known_spans):
            continue
        label = re.sub(r"\s+", " ", match.group("label")).strip(" -:;,.()").casefold()
        unit = (match.group("unit") or "").strip()
        if not LAB_HINT_RE.search(label):
            continue
        if re.search(r"(?:diameter|dimension|ejection fraction|magnification|body mass index)$", label):
            continue
        output.append((label, float(match.group("value").replace(",", "")), unit))
    return output


def build_report(
    records: Iterable[TextRecord],
    schema_checks: dict[str, dict],
    dataset_manifest: dict,
    source_manifest: dict,
    extension: dict,
    catalog: dict,
    normalizer: FindingNormalizer,
) -> dict:
    patterns = _alias_patterns(catalog)
    per_test: dict[str, dict] = collections.defaultdict(
        lambda: {
            "occurrences": 0,
            "case_keys": set(),
            "datasets": collections.Counter(),
            "units": collections.Counter(),
            "unsupported_units": collections.Counter(),
            "unit_missing_occurrences": 0,
            "context_required_occurrences": 0,
            "inline_reference_occurrences": 0,
            "examples": [],
        }
    )
    per_dataset: dict[str, dict] = collections.defaultdict(
        lambda: {
            "rows_with_measurements": set(),
            "measurements": 0,
            "tests": set(),
            "inline_reference_occurrences": 0,
            "unsupported_unit_occurrences": 0,
            "unit_missing_occurrences": 0,
            "context_required_occurrences": 0,
        }
    )
    unresolved: dict[str, dict] = collections.defaultdict(
        lambda: {"occurrences": 0, "datasets": collections.Counter(), "units": collections.Counter(), "examples": []}
    )

    for record in records:
        measurements = scan_measurements(record.text, patterns)
        known_spans = [(m.start, m.end) for m in measurements]
        for measurement in measurements:
            stats = per_test[measurement.canonical]
            dataset_stats = per_dataset[record.dataset_id]
            key = f"{record.dataset_id}:{record.row_id}"
            stats["occurrences"] += 1
            stats["case_keys"].add(key)
            stats["datasets"][record.dataset_id] += 1
            stats["units"][measurement.unit or "<unitless>"] += 1
            dataset_stats["rows_with_measurements"].add(record.row_id)
            dataset_stats["measurements"] += 1
            dataset_stats["tests"].add(measurement.canonical)

            atom = f"{measurement.alias}: {measurement.qualifier}{measurement.value:g} {measurement.unit}".strip()
            if len(stats["examples"]) < 3 and atom not in stats["examples"]:
                stats["examples"].append(atom)
            if normalizer._extract_inline_reference(record.text[measurement.start : measurement.end + 120], measurement.unit):
                stats["inline_reference_occurrences"] += 1
                dataset_stats["inline_reference_occurrences"] += 1
            unit_status = _unit_status(normalizer, catalog, measurement)
            if unit_status == "unsupported_unit":
                stats["unsupported_units"][measurement.unit or "<unitless>"] += 1
                dataset_stats["unsupported_unit_occurrences"] += 1
            elif unit_status == "unit_missing":
                stats["unit_missing_occurrences"] += 1
                dataset_stats["unit_missing_occurrences"] += 1
            elif unit_status == "context_required":
                stats["context_required_occurrences"] += 1
                dataset_stats["context_required_occurrences"] += 1

        for label, value, unit in _generic_candidates(record.text, known_spans):
            if normalizer._resolve_std_name(label) is not None:
                continue
            stats = unresolved[label]
            stats["occurrences"] += 1
            stats["datasets"][record.dataset_id] += 1
            stats["units"][unit or "<unitless>"] += 1
            atom = f"{label}: {value:g} {unit}".strip()
            if len(stats["examples"]) < 2 and atom not in stats["examples"]:
                stats["examples"].append(atom)

    tests = []
    for canonical, stats in per_test.items():
        entry = catalog[canonical]
        tests.append(
            {
                "test": canonical,
                "occurrences": stats["occurrences"],
                "cases": len(stats["case_keys"]),
                "datasets": dict(sorted(stats["datasets"].items())),
                "observed_units": dict(stats["units"].most_common()),
                "unsupported_units": dict(stats["unsupported_units"].most_common()),
                "unit_missing_occurrences": stats["unit_missing_occurrences"],
                "context_required_occurrences": stats["context_required_occurrences"],
                "inline_reference_occurrences": stats["inline_reference_occurrences"],
                "reference_status": "static_fallback" if entry.get("reference_ranges") else "context_or_local_reference_required",
                "reference_sources": _reference_sources(entry),
                "reference_ranges": entry.get("reference_ranges", []),
                "decision_limits": entry.get("decision_limits", []),
                "caution": entry.get("caution"),
                "examples": stats["examples"],
            }
        )
    tests.sort(key=lambda item: (-item["occurrences"], item["test"]))

    unresolved_items = [
        {
            "label": label,
            "occurrences": stats["occurrences"],
            "datasets": dict(sorted(stats["datasets"].items())),
            "observed_units": dict(stats["units"].most_common()),
            "examples": stats["examples"],
            "status": "human_review_required",
        }
        for label, stats in unresolved.items()
        if stats["occurrences"] >= 1
    ]
    unresolved_items.sort(key=lambda item: (-item["occurrences"], item["label"]))

    dataset_results = []
    for dataset in dataset_manifest["datasets"]:
        if not dataset.get("numeric_lab_audit"):
            dataset_results.append(
                {
                    "id": dataset["id"],
                    "rows": dataset["rows"],
                    "numeric_lab_audit": False,
                    "exclusion_reason": dataset.get("exclusion_reason"),
                    "schema_check": schema_checks.get(dataset["id"]),
                }
            )
            continue
        stats = per_dataset[dataset["id"]]
        dataset_results.append(
            {
                "id": dataset["id"],
                "rows": dataset["rows"],
                "numeric_lab_audit": True,
                "rows_with_catalog_measurements": len(stats["rows_with_measurements"]),
                "catalog_measurements": stats["measurements"],
                "distinct_catalog_tests": len(stats["tests"]),
                "inline_reference_occurrences": stats["inline_reference_occurrences"],
                "unsupported_unit_occurrences": stats["unsupported_unit_occurrences"],
                "unit_missing_occurrences": stats["unit_missing_occurrences"],
                "context_required_occurrences": stats["context_required_occurrences"],
            }
        )

    return {
        "schema_version": "1.0",
        "generated_on": dataset_manifest.get("generated_on"),
        "policy": source_manifest["interpretation_policy"],
        "dataset_manifest": "data/eval/lab_reference_dataset_manifest.json",
        "source_manifest": "data/knowledge_raw/lab_reference_sources.json",
        "catalog_summary": {
            "total_tests": len(catalog),
            "legacy_tests": len(catalog) - len(extension.get("additions", {})),
            "extension_additions": len(extension.get("additions", {})),
            "detected_tests": len(tests),
            "detected_measurements": sum(item["occurrences"] for item in tests),
            "cases_with_measurements": len({key for stats in per_test.values() for key in stats["case_keys"]}),
            "tests_requiring_context_or_local_reference": sum(
                item["reference_status"] != "static_fallback" for item in tests
            ),
            "unsupported_unit_occurrences": sum(
                sum(item["unsupported_units"].values()) for item in tests
            ),
            "unit_missing_occurrences": sum(item["unit_missing_occurrences"] for item in tests),
            "context_required_occurrences": sum(item["context_required_occurrences"] for item in tests),
        },
        "datasets": dataset_results,
        "detected_tests": tests,
        "uncatalogued_candidates": unresolved_items[:100],
        "limitations": [
            "This is a conservative regex and alias audit, not a clinical NLP gold standard; prose-only and unusually formatted measurements may be missed.",
            "An unsupported unit can be a dataset extraction error, a method-specific unit, or a missing conversion; it is never numerically compared by the runtime.",
            "Tumor markers, high-sensitivity troponin, D-dimer, reproductive hormones, and NT-proBNP require method or clinical context even when representative limits are listed.",
            "Raw benchmark files are checksum-verified locally and are not redistributed by this repository."
        ],
    }


def _fmt_ranges(test: dict) -> str:
    records = test.get("reference_ranges") or test.get("decision_limits") or []
    output = []
    for record in records[:3]:
        if "value" in record:
            output.append(f"{record.get('operator', '')}{record['value']} {record.get('unit', '')}".strip())
        else:
            low, high = record.get("low"), record.get("high")
            if low is None:
                span = f"≤{high}"
            elif high is None:
                span = f"≥{low}"
            else:
                span = f"{low}–{high}"
            context = record.get("gender") or record.get("population")
            output.append(f"{span} {record.get('unit', '')}{f' ({context})' if context and context != 'any' else ''}".strip())
    return "; ".join(output) if output else "local/context required"


def render_markdown(report: dict, sources: dict, datasets: dict) -> str:
    summary = report["catalog_summary"]
    lines = [
        "# Lab reference-range extension and benchmark audit",
        "",
        f"> Generated {report['generated_on']}. Research/benchmark normalization only; not for patient care.",
        "",
        "## Outcome",
        "",
        f"The catalog expands from **{summary['legacy_tests']} to {summary['total_tests']} tests** "
        f"(**{summary['extension_additions']} additions**). The pinned numeric-bearing splits produced "
        f"**{summary['detected_measurements']} catalog measurements across {summary['cases_with_measurements']} cases**, "
        f"covering **{summary['detected_tests']} distinct tests**.",
        "",
        "The runtime applies an interval printed in the case first. A static interval is used only when the unit and available context match; incompatible units return `unknown`. FEU and DDU are never interconverted.",
        "",
        "## Target datasets",
        "",
        "| Dataset | Rows | Numeric audit | Measurements | Distinct tests | Notes |",
        "|---|---:|:---:|---:|---:|---|",
    ]
    by_id = {item["id"]: item for item in datasets["datasets"]}
    for item in report["datasets"]:
        manifest_item = by_id[item["id"]]
        if item["numeric_lab_audit"]:
            lines.append(
                f"| [{manifest_item['name']}]({manifest_item['source_page']}) | {item['rows']} | yes | "
                f"{item['catalog_measurements']} | {item['distinct_catalog_tests']} | "
                f"{item['unsupported_unit_occurrences']} incompatible units; "
                f"{item['unit_missing_occurrences']} unitless; "
                f"{item['context_required_occurrences']} context-dependent |"
            )
        else:
            lines.append(
                f"| [{manifest_item['name']}]({manifest_item['source_page']}) | {item['rows']} | no | — | — | "
                f"{item['exclusion_reason']} |"
            )

    lines.extend(
        [
            "",
            "## Detected tests and fallback values",
            "",
            "The table is occurrence-ranked. Values are representative benchmark fallbacks, not replacements for the reporting laboratory interval.",
            "The versioned 2026/Mayo/NHS source IDs support this extension; older source IDs are legacy values retained from the pre-existing catalog, as recorded in the source manifest.",
            "",
            "| Test | Occurrences / cases | Observed units | Fallback or decision limit | Source |",
            "|---|---:|---|---|---|",
        ]
    )
    for test in report["detected_tests"]:
        units = ", ".join(f"{unit} ({count})" for unit, count in test["observed_units"].items())
        source_ids = ", ".join(test["reference_sources"]) or "local/context"
        lines.append(
            f"| `{test['test']}` | {test['occurrences']} / {test['cases']} | {units} | "
            f"{_fmt_ranges(test)} | {source_ids} |"
        )

    lines.extend(
        [
            "",
            "## Reference sources",
            "",
            "| ID | Source | Version/use |",
            "|---|---|---|",
        ]
    )
    for source_id, source in sources["sources"].items():
        version = source.get("version", "")
        lines.append(
            f"| `{source_id}` | [{source['title']}]({source['url']}) | "
            f"{version + '; ' if version else ''}{source['use']} |"
        )

    lines.extend(
        [
            "",
            "## Reproduction",
            "",
            "```bash",
            "pip install -e '.[lab-audit]'",
            "python scripts/download_lab_audit_datasets.py",
            "python scripts/extend_lab_reference_data.py --check",
            "python scripts/audit_lab_reference_coverage.py",
            "```",
            "",
            "The downloader pins every artifact by revision, byte count, and SHA-256. Raw datasets remain local because their redistribution terms differ.",
            "",
            "## Safety and limitations",
            "",
        ]
    )
    lines.extend(f"- {item}" for item in report["limitations"])
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset-dir", type=Path, default=DEFAULT_DATASET_DIR)
    parser.add_argument("--dataset-manifest", type=Path, default=DEFAULT_DATASET_MANIFEST)
    parser.add_argument("--source-manifest", type=Path, default=DEFAULT_SOURCE_MANIFEST)
    parser.add_argument("--extension", type=Path, default=DEFAULT_EXTENSION)
    parser.add_argument("--catalog", type=Path, default=DEFAULT_CATALOG)
    parser.add_argument("--loinc2hpo", type=Path, default=DEFAULT_LOINC)
    parser.add_argument("--conversions", type=Path, default=DEFAULT_CONVERSIONS)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--output-markdown", type=Path, default=DEFAULT_MARKDOWN)
    args = parser.parse_args()

    dataset_manifest = json.loads(args.dataset_manifest.read_text(encoding="utf-8"))
    source_manifest = json.loads(args.source_manifest.read_text(encoding="utf-8"))
    extension = json.loads(args.extension.read_text(encoding="utf-8"))
    catalog = json.loads(args.catalog.read_text(encoding="utf-8"))
    normalizer = FindingNormalizer(args.catalog, args.loinc2hpo, args.conversions)
    records, schema_checks = iter_dataset_texts(args.dataset_dir, dataset_manifest)
    report = build_report(
        records,
        schema_checks,
        dataset_manifest,
        source_manifest,
        extension,
        catalog,
        normalizer,
    )

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    args.output_markdown.write_text(
        render_markdown(report, source_manifest, dataset_manifest), encoding="utf-8"
    )
    print(
        f"WROTE {args.output_json.relative_to(ROOT)} and {args.output_markdown.relative_to(ROOT)}: "
        f"{report['catalog_summary']['detected_measurements']} measurements, "
        f"{report['catalog_summary']['detected_tests']} tests"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
